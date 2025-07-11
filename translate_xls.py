from openpyxl import Workbook
from googletrans import Translator, constants
import asyncio

async def translate_text(text, lang):
    translator = Translator()
    
    return translation.text

# Create a new workbook in English
we = Workbook()

# Get the active worksheet
ws = we.active

# Add a new worksheet at the end (default)
# ws1 = we.create_sheet("Inventory")


sheetName = "Inventory"
titles = ["Item", "Quantity", "Price"]
items = ["Apples", "Bananas", "Cherries"]

ws.title = sheetName
# Write a table out to the worksheet
ws["A1"] = titles[0]
ws["B1"] = titles[1]
ws["C1"] = titles[2]    

ws["A2"] = items[0]
ws["B2"] = 10
ws["C2"] = 0.5
ws["A3"] = items[1]
ws["B3"] = 20
ws["C3"] = 0.25
ws["A4"] = items[2]
ws["B4"] = 15  
ws["C4"] = 0.75

we.save('englishworkbook.xlsx')

# Translate the workbook to French

sheetName = asyncio.run(translate_text(sheetName, 'fr'))
titles = [asyncio.run(translate_text(title, 'fr')) for title in titles] 
items =  [asyncio.run(translate_text(item, 'fr')) for item in items]

wf = Workbook()
ws = wf.active

ws.title = sheetName
ws["A1"] = titles[0]
ws["B1"] = titles[1]
ws["C1"] = titles[2] 
ws["A2"] = items[0]
ws["B2"] = 10
ws["C2"] = 0.5
ws["A3"] = items[1]
ws["B3"] = 20
ws["C3"] = 0.25
ws["A4"] = items[2]
ws["B4"] = 15  
ws["C4"] = 0.75

wf.save('frenchworkbook.xlsx')